import os
import json
import hashlib
from pathlib import Path
from bs4 import BeautifulSoup
import docx
import openpyxl
from PyPDF2 import PdfReader

class FolderIndexer:
    """Handles indexing of files in specified folders into a ChromaDB collection."""

    def __init__(self, collection, db):
        self.collection = collection
        self.db = db

    def _calculate_file_hash(self, filepath):
        """Calculates the SHA-256 hash of a file."""
        hasher = hashlib.sha256()
        try:
            with open(filepath, 'rb') as f:
                while chunk := f.read(8192):
                    hasher.update(chunk)
            return hasher.hexdigest()
        except (IOError, FileNotFoundError) as e:
            print(f"Error calculating hash for {filepath}: {e}")
            return None

    def _split_text(self, text, chunk_size=1000, chunk_overlap=200):
        """A simple text splitter."""
        if not text:
            return []
        
        chunks = []
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunks.append(text[start:end])
            start += chunk_size - chunk_overlap
            if start < 0: # overlap is larger than chunk size
                start = end
        return chunks

    def _extract_text(self, filepath):
        """Extracts text content from a file based on its extension."""
        content = ""
        try:
            extension = os.path.splitext(filepath)[1].lower()
            if extension in ['.txt', '.md', '.json']:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
            elif extension in ['.html', '.htm']:
                with open(filepath, 'r', encoding='utf-8') as f:
                    soup = BeautifulSoup(f, 'html.parser')
                    content = soup.get_text(separator='\n', strip=True)
            elif extension == '.docx':
                doc = docx.Document(filepath)
                content = '\n'.join([para.text for para in doc.paragraphs])
            elif extension == '.xlsx':
                workbook = openpyxl.load_workbook(filepath, read_only=True)
                full_text = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        row_text = [str(cell.value) for cell in row if cell.value]
                        if row_text:
                            full_text.append(' '.join(row_text))
                content = '\n'.join(full_text)
            elif extension == '.pdf':
                with open(filepath, 'rb') as f:
                    reader = PdfReader(f)
                    full_text = [page.extract_text() for page in reader.pages if page.extract_text()]
                    content = '\n'.join(full_text)
        except Exception as e:
            print(f"Error extracting text from {filepath}: {e}")
            return None
        return content

    def index_folder(self, folder_path_str):
        """Indexes files in a folder, handling new, modified, and deleted files."""
        print(f"Indexing folder: {folder_path_str}")
        folder_path = Path(folder_path_str)
        if not folder_path.is_dir():
            print(f"Error: {folder_path_str} is not a valid directory.")
            return

        indexed_files_db = self.db.get_indexed_files_by_folder(folder_path_str)
        
        current_files_on_disk = set()
        files_to_index = []

        for filepath in folder_path.rglob('*'):
            if filepath.is_file() and not filepath.name.startswith('.'):
                str_filepath = str(filepath)
                current_files_on_disk.add(str_filepath)
                
                try:
                    mtime = os.path.getmtime(filepath)
                    size = os.path.getsize(filepath)
                    
                    db_record = indexed_files_db.get(str_filepath)

                    if not db_record or mtime > db_record['modified_time'] or size != db_record['size']:
                        file_hash = self._calculate_file_hash(filepath)
                        if not file_hash: continue

                        if not db_record or file_hash != db_record['file_hash']:
                            files_to_index.append({
                                "path": filepath,
                                "mtime": mtime,
                                "size": size,
                                "hash": file_hash
                            })

                except FileNotFoundError:
                    continue

        indexed_files_paths_db = set(indexed_files_db.keys())
        deleted_files_paths = indexed_files_paths_db - current_files_on_disk
        
        if deleted_files_paths:
            ids_to_delete_chroma = []
            for str_filepath in deleted_files_paths:
                db_record = indexed_files_db.get(str_filepath)
                if db_record and db_record.get('chunk_ids'):
                    chunk_ids = json.loads(db_record['chunk_ids'])
                    ids_to_delete_chroma.extend(chunk_ids)
                self.db.delete_indexed_file(str_filepath)
            
            if ids_to_delete_chroma:
                self.collection.delete(ids=ids_to_delete_chroma)
            print(f"Removed {len(deleted_files_paths)} deleted files from index.")

        for file_info in files_to_index:
            filepath = file_info['path']
            str_filepath = str(filepath)
            print(f"Processing: {str_filepath}")
            
            # If file was already indexed, delete old chunks from Chroma
            db_record = indexed_files_db.get(str_filepath)
            if db_record and db_record.get('chunk_ids'):
                old_chunk_ids = json.loads(db_record['chunk_ids'])
                if old_chunk_ids:
                    self.collection.delete(ids=old_chunk_ids)

            content = self._extract_text(filepath)
            if content:
                chunks = self._split_text(content)
                chunk_ids = [f"{str_filepath}_{i}" for i in range(len(chunks))]
                
                if chunks:
                    batch_size = 1000  # A safe batch size to avoid exceeding limits
                    for i in range(0, len(chunks), batch_size):
                        batch_chunks = chunks[i:i + batch_size]
                        batch_ids = chunk_ids[i:i + batch_size]
                        batch_metadatas = [{"source": str_filepath}] * len(batch_chunks)
                        
                        self.collection.add(
                            documents=batch_chunks,
                            metadatas=batch_metadatas,
                            ids=batch_ids
                        )
                
                # Update database with new file info
                self.db.update_indexed_file(
                    filepath=str_filepath,
                    size=file_info['size'],
                    modified_time=file_info['mtime'],
                    file_hash=file_info['hash'],
                    chunk_ids=chunk_ids
                )
        
        print(f"Finished indexing {folder_path_str}.")

    def remove_folder_from_index(self, folder_path_str):
        """Removes all indexed files for a given folder from ChromaDB and the database."""
        print(f"Removing folder {folder_path_str} from index...")
        indexed_files_db = self.db.get_indexed_files_by_folder(folder_path_str)
        
        if not indexed_files_db:
            print("Folder was not in the index.")
            return

        ids_to_delete_chroma = []
        for str_filepath, db_record in indexed_files_db.items():
            if db_record and db_record.get('chunk_ids'):
                chunk_ids = json.loads(db_record['chunk_ids'])
                ids_to_delete_chroma.extend(chunk_ids)
            
            # Delete from SQLite DB
            self.db.delete_indexed_file(str_filepath)

        if ids_to_delete_chroma:
            self.collection.delete(ids=ids_to_delete_chroma)
        
        print(f"Removed folder {folder_path_str} and {len(indexed_files_db)} files from index.")
