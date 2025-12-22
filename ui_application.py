# [file name]: ui_application.py (updated)
"""UI application and device management for recass."""

import os
import threading
import queue
import time
from datetime import datetime, timedelta
import gi
gi.require_version('Gtk', '3.0')
from gi.repository import Gtk, GLib
import torch
import whisper
from pyannote.audio import Pipeline

from collections import deque

from audio_recorder import AudioRecorder
from transcriber import Transcriber
from screenshot_manager import ScreenshotManager
from config import load_user_settings, save_user_settings
import config as cfg
from database import Database
from folder_indexer import FolderIndexer

# Import the new modules
from system_tray import SystemTrayManager
from chat_window import ChatWindow
from timer_manager import TimerManager
from joplin_sync import JoplinSync
from audio_device_manager import AudioDeviceManager
from folder_manager import FolderManager
from recording_window import RecordingIndicatorWindow
from history_window import HistoryWindow
from ollama_analyzer import OllamaAnalyzer
from meeting_browser_window import MeetingBrowserWindow # New Import
from chat_browser_window import ChatBrowserWindow # New Import


class Application:
    """Main application handling UI, device management, and file recording."""

    def __init__(self):
        # Initialize core components
        self.recorder = None
        self.transcriber = None
        self.whisper_model = None
        self.diarization_pipeline = None
        self.screenshot_manager = None
        self.db = Database()
        self.folder_indexer = None
        
        # Initialize new modular components
        self.system_tray = SystemTrayManager(self)
        self.chat_window = ChatWindow(self)
        self.timer_manager = TimerManager(self)
        self.joplin_sync = JoplinSync(self)
        self.audio_device_manager = AudioDeviceManager(self)
        self.folder_manager = FolderManager(self)
        self.meeting_browser_window = None # New initialization
        self.chat_browser_window = None # New initialization
        
        # Core application state
        self.is_recording = False
        self.is_starting = False
        self.output_file = None
        self.output_filename = None
        self.file_lock = threading.Lock()
        self.meeting_folder = None
        self.chat_session_id = None
        self.current_meeting_id = None # New attribute initialization
        self.consistency_check_thread = None
        self.consistency_check_stop_event = threading.Event()
        self.consistency_check_enabled = False
        
        # Device IDs
        self.mic_dev_id = None
        self.loopback_dev_id = None
        
        # UI references
        self._window = None
        self.mic_combo = None
        self.loopback_combo = None
        self.loopback_level_bar = None
        self._record_button = None
        self.folders_listbox = None
        self.recording_window = None
        self.history_window = None
        self.transcription_history = []
        self.recent_transcriptions = deque() # Stores (timestamp, line) for the last 5 minutes
        
        # Load settings
        self._load_settings()
        
        # Language choices
        self._all_lang_choices = [
            ("Auto (detect)", "auto"),
            ("English (en)", "en"),
            ("Deutsch (de)", "de"),
            ("Español (es)", "es"),
            ("Français (fr)", "fr"),
            ("Italiano (it)", "it"),
            ("Português (pt)", "pt"),
            ("Nederlands (nl)", "nl"),
            ("中文 / Chinese (zh)", "zh"),
            ("日本語 / Japanese (ja)", "ja"),
            ("한국어 / Korean (ko)", "ko"),
            ("Русский / Russian (ru)", "ru"),
            ("العربية / Arabic (ar)", "ar"),
            ("हिन्दी / Hindi (hi)", "hi"),
            ("Türkçe / Turkish (tr)", "tr"),
            ("Polski / Polish (pl)", "pl"),
            ("Svenska / Swedish (sv)", "sv"),
            ("Norsk / Norwegian (no)", "no"),
            ("Dansk / Danish (da)", "da"),
            ("Suomi / Finnish (fi)", "fi"),
            ("Čeština / Czech (cs)", "cs"),
            ("Ελληνικά / Greek (el)", "el"),
            ("Magyar / Hungarian (hu)", "hu"),
            ("Română / Romanian (ro)", "ro"),
            ("Українська / Ukrainian (uk)", "uk"),
            ("Bahasa Indonesia (id)", "id"),
            ("Bahasa Melayu (ms)", "ms"),
            ("Tiếng Việt / Vietnamese (vi)", "vi"),
        ]
        self._lang_code_map = [code for _, code in self._all_lang_choices]
    
    def _load_settings(self):
        """Load all settings from config."""
        settings = load_user_settings()
        
        # Screenshot settings
        self.screenshot_disabled = bool(settings.get('screenshot_disabled', False))
        self.screenshot_target = settings.get('screenshot_target', 'all')
        self.screenshot_interval = int(settings.get('screenshot_interval', 10))
        self.send_screenshots_to_llm = bool(settings.get('send_screenshots_to_llm', True))
        
        # AI settings
        self.ai_record_meeting = bool(settings.get('ai_record_meeting', True))
        
        # Ollama settings
        self.ollama_model_name = settings.get('ollama_model_name', 'llama3')
        
        # Whisper settings
        self.whisper_model_name = settings.get('whisper_model_name', 'base')
        
        # Timer settings
        self.timer_enabled = bool(settings.get('timer_enabled', False))
        self.timer_setting_str = settings.get('timer_value', '00:00:00')
        if not self.timer_setting_str:
            self.timer_setting_str = '00:00:00'
        self.timer_manager.timer_seconds = (self.timer_manager.hms_to_seconds(
            self.timer_setting_str) or 0)
        self.timer_manager.remaining_seconds = self.timer_manager.timer_seconds
        self.timer_manager.timer_enabled = self.timer_enabled
        self.timer_manager.timer_setting_str = self.timer_setting_str
        
        # Source folders
        self.source_folders = settings.get('source_folders', [])
        
        # Joplin settings
        self.joplin_api_key = settings.get('joplin_api_key', '')
        self.joplin_sync_enabled = bool(settings.get('joplin_sync_enabled', False))
        self.joplin_destination_folder = settings.get('joplin_destination_folder', '')
        
        # Consistency check setting
        self.consistency_check_enabled = bool(settings.get('consistency_check_enabled', False))
        
        # Device names
        self.audio_device_manager.mic_dev_name = settings.get('mic_dev_name')
        self.audio_device_manager.loopback_dev_name = settings.get('loopback_dev_name')

    def _load_diarization_pipeline(self):
        """Load the speaker diarization pipeline if HF token is available."""
        hf_token = os.environ.get("HUGGING_FACE_TOKEN") or (cfg.HF_TOKEN if hasattr(cfg, 'HF_TOKEN') else None)
        if not hf_token:
            print("⚠️ WARNUNG: HUGGING_FACE_TOKEN nicht gefunden. Speaker Diarization für COMP-Quelle ist deaktiviert.")
            return None

        print("🧠 Lade Pyannote Diarization-Pipeline... Dies kann einen Moment dauern.")
        
        # Monkey-patch torch.load to set weights_only=False as requested for this specific model.
        original_torch_load = torch.load
        def unsafe_torch_load(*args, **kwargs):
            kwargs['weights_only'] = False
            return original_torch_load(*args, **kwargs)
        torch.load = unsafe_torch_load
        
        try:
            print("  → Laden von 'pyannote/speaker-diarization-3.1'...")
            pipeline = Pipeline.from_pretrained(
                "pyannote/speaker-diarization-3.1",
                use_auth_token=hf_token
            )
            print("  ✓ Pipeline-Modell geladen")
            
            # Sende die Pipeline auf die GPU, falls verfügbar
            if torch.cuda.is_available():
                try:
                    print("  → GPU verfügbar, verschiebe Pipeline zu CUDA...")
                    pipeline.to(torch.device("cuda"))
                    print("  ✓ Pipeline auf GPU geladen")
                except torch.OutOfMemoryError:
                    print("  ⚠️ CUDA out of memory. Lasse Diarization-Pipeline auf CPU.")
                except Exception as e:
                    print(f"  ⚠️ Fehler beim Verschieben der Pipeline auf GPU: {e}. Lasse auf CPU.")
            else:
                print("  → GPU nicht verfügbar, nutze CPU")
            
            print("✅ Diarization-Pipeline vollständig geladen und bereit.")
            return pipeline
        except Exception as e:
            print(f"❌ Fehler beim Laden der Diarization-Pipeline: {e}")
            import traceback
            print(f"   Traceback: {traceback.format_exc()}")
            print("   Stellen Sie sicher, dass Sie die User Conditions auf Hugging Face akzeptiert haben und Ihr Token korrekt ist.")
            return None
        finally:
            # Restore original torch.load to prevent side effects
            torch.load = original_torch_load
    
    def _run_audio_processing_loop(self):
        """Main audio processing loop running in separate thread."""
        try:
            # Ensure transcriber only processes real-time audio when recording is active
            try:
                self.transcriber.recording_enabled = bool(self.is_recording)
            except Exception:
                pass
            
            try:
                self.recorder.start_recording()
                print("DEBUG: recorder.start_recording() called successfully.")
            except Exception as e:
                print(f"ERROR: recorder.start_recording() failed: {e}")
                GLib.idle_add(self._show_error_dialog, "Audio Error", f"Could not start recording. Please check your audio devices.\n\n{e}")
                GLib.idle_add(self._stop_file_recording)
                return

            print("\n--- Aufnahme und Transkription läuft. Icon in der Taskleiste. ---\n")
            
            while self.transcriber and self.transcriber.is_alive():
                self.transcriber.join(timeout=0.5)
        except Exception as e:
            print(f"\nEin unerwarteter Fehler im Audio-Thread ist aufgetreten: {e}")
            GLib.idle_add(self._show_error_dialog, "Unexpected Audio Error", f"An unexpected error occurred in the audio thread.\n\n{e}")
            GLib.idle_add(self.quit_action, None, None)

    def _run_consistency_check_loop(self):
        """
        Periodically checks for inconsistencies using Ollama if the consistency check is enabled
        and recording is active. Adds notes to meeting minutes if inconsistencies are found.
        """
        check_interval = 60 # seconds
        while not self.consistency_check_stop_event.is_set():
            if self.is_recording and self.consistency_check_enabled:
                # Get recent transcriptions
                recent_history_lines = [item[1] for item in self.recent_transcriptions]
                current_meeting_history = "\n".join(recent_history_lines)

                if current_meeting_history.strip():
                    # Fetch relevant past meeting context
                    previous_meeting_context = self._get_relevant_past_meeting_context(current_meeting_history)

                    if previous_meeting_context.strip():
                        analyzer = self._get_ollama_analyzer()
                        inconsistencies = analyzer.find_inconsistencies(current_meeting_history, previous_meeting_context)

                        if inconsistencies.strip():
                            note = f"\n[CONSISTENCY CHECK NOTE]: {inconsistencies}\n"
                            self._add_note_to_meeting_minutes(note)
                            if self.history_window and self.history_window.is_visible():
                                GLib.idle_add(self.history_window.append_text, note)            
            # Wait for interval or until stop event is set
            self.consistency_check_stop_event.wait(check_interval)
    
    def stop_threads(self):
        """Stop recorder and transcriber threads."""
        if self.recorder:
            print("Beende Aufnahme...")
            self.recorder.stop_recording()
            self.recorder = None
        if self.transcriber and self.transcriber.is_alive():
            print("Beende Transkription...")
            self.transcriber.stop()
            self.transcriber.join()
            self.transcriber = None
    
    def _start_audio_processing_thread(self, mic_id: int, loopback_id: int):
        """Starts or restarts the audio processing thread with specified device IDs."""
        print(f"DEBUG: _start_audio_processing_thread called with mic_id={mic_id}, loopback_id={loopback_id}")
        
        # Stop any existing threads
        self.stop_threads()
        
        # Update instance variables
        self.mic_dev_id = mic_id
        self.loopback_dev_id = loopback_id
        
        # Initialize recorder and transcriber synchronously here
        transcription_queue = queue.Queue()
        self.transcriber = Transcriber(
            self.whisper_model,
            self.diarization_pipeline,
            transcription_queue, 
            self._on_new_transcription,
            self._on_analysis_complete
        )
        self.transcriber.start() # Start transcriber thread immediately

        print(f"DEBUG: Creating AudioRecorder with mic_dev_id={self.mic_dev_id} and loopback_dev_id={self.loopback_dev_id}")
        self.recorder = AudioRecorder(
            self.mic_dev_id, 
            self.loopback_dev_id, 
            transcription_queue,
            level_callback=self._on_level_update
        )
        print("DEBUG: AudioRecorder instance created.")
        
        # Start a new thread for the audio processing loop
        self._app_thread = threading.Thread(target=self._run_audio_processing_loop)
        self._app_thread.daemon = True
        self._app_thread.start()
        print("DEBUG: Audio processing thread started.")
    
    # ==================== UI Creation Methods ====================
    
    def _show_error_dialog(self, primary_text, secondary_text):
        """Shows a GTK error message dialog."""
        dialog = Gtk.MessageDialog(
            transient_for=self._window,
            flags=0,
            message_type=Gtk.MessageType.ERROR,
            buttons=Gtk.ButtonsType.CANCEL,
            text=primary_text,
        )
        dialog.format_secondary_text(secondary_text)
        dialog.run()
        dialog.destroy()
    
    def _create_or_show_window(self):
        """Create or show the main GTK window."""
        if self._window:
            self._window.present()
            return
        
        self._window = Gtk.Window(title="recass Status")
        self._window.set_resizable(False)
        self._window.connect("delete-event", lambda w, e: w.hide() or True)
        
        grid = Gtk.Grid(margin=10, column_spacing=10, row_spacing=10)
        self._window.add(grid)
        
        # Device Selection
        l_mic = Gtk.Label(label="Microphone:")
        l_mic.set_xalign(0)
        self.mic_combo = Gtk.ComboBoxText()
        self.mic_combo.set_size_request(400, -1)
        self.mic_combo.set_sensitive(not self.is_recording)
        
        l_loop = Gtk.Label(label="Computer-Audio:")
        l_loop.set_xalign(0)
        self.loopback_combo = Gtk.ComboBoxText()
        self.loopback_combo.set_sensitive(not self.is_recording)

        # Level bar for loopback
        self.loopback_level_bar = Gtk.LevelBar()
        self.loopback_level_bar.set_min_value(0)
        self.loopback_level_bar.set_max_value(0.1) # Initial guess, may need tuning
        self.loopback_level_bar.set_value(0)
        
        loopback_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        loopback_box.pack_start(self.loopback_combo, True, True, 0)
        loopback_box.pack_start(self.loopback_level_bar, True, True, 0)

        self.audio_device_manager.populate_audio_devices()
        
        self.mic_combo_handler_id = self.mic_combo.connect("changed", self.audio_device_manager.on_device_changed)
        self.loopback_combo_handler_id = self.loopback_combo.connect("changed", self.audio_device_manager.on_device_changed)
        
        grid.attach(l_mic, 0, 0, 1, 1)
        grid.attach(self.mic_combo, 1, 0, 1, 1)
        grid.attach(l_loop, 0, 1, 1, 1)
        grid.attach(loopback_box, 1, 1, 2, 1)
        
        self.audio_refresh_btn = Gtk.Button(label="Refresh")
        self.audio_refresh_btn.connect("clicked", self.audio_device_manager.populate_audio_devices)
        self.audio_refresh_btn.set_sensitive(not self.is_recording)
        grid.attach(self.audio_refresh_btn, 2, 0, 1, 1)
        
        # Screenshot selection
        self._create_screenshot_ui(grid, 2)
        
        # Screenshot interval
        self._create_screenshot_interval_ui(grid, 3)

        # Send screenshots to LLM toggle
        self._create_send_screenshots_to_llm_ui(grid, 4)
        
        # Consistency check toggle
        self._create_consistency_check_ui(grid, 4)
        
        # Diarization Speaker Count Controls
        self._create_speaker_controls(grid, 5)
        
        # Transcription language selection
        self._create_language_ui(grid, 6)
        
        # AI recording toggle
        self._create_ai_ui(grid, 6)
        
        # Ollama URL setting
        self._create_ollama_ui(grid, 7)
        
        # Hugging Face token setting
        self._create_hf_token_ui(grid, 9)
        
        # Whisper Model setting
        self._create_whisper_model_ui(grid, 10)
        
        # Joplin API Key setting
        self._create_joplin_ui(grid, 11)
        
        # Source Folders ListBox
        self._create_folder_ui(grid, 14)
        
        # Timer settings
        self._create_timer_ui(grid, 15)
        
        # Action Buttons
        self._create_action_buttons(grid, 16)
        
        self._window.show_all()
    
    def _create_whisper_model_ui(self, grid, row):
        """Create Whisper model setting UI."""
        l_whisper = Gtk.Label(label="Whisper Model:")
        l_whisper.set_xalign(0)
        self.whisper_entry = Gtk.Entry()
        try:
            settings = load_user_settings()
            self.whisper_entry.set_text(settings.get('whisper_model_name', 'base'))
        except Exception:
            self.whisper_entry.set_text('base')
        self.whisper_entry.set_size_request(400, -1)
        self.whisper_entry.connect("changed", self._on_whisper_model_changed)
        grid.attach(l_whisper, 0, row, 1, 1)
        grid.attach(self.whisper_entry, 1, row, 2, 1)
        
    def _create_screenshot_ui(self, grid, row):
        """Create screenshot-related UI elements."""
        l_screenshot = Gtk.Label(label="Screen to capture:")
        l_screenshot.set_xalign(0)
        self.screenshot_disable_btn = Gtk.CheckButton(label="Disable screen capture")
        self.screenshot_disable_btn.set_active(bool(self.screenshot_disabled))
        self.screenshot_disable_btn.connect("toggled", self._on_screenshot_disabled_toggled)
        
        self.screenshot_combo = Gtk.ComboBoxText()
        self.screenshot_combo.set_size_request(400, -1)
        
        try:
            targets = ScreenshotManager.list_capture_screens()
            self._screenshot_target_map = [tid for tid, _ in targets]
            
            for i, (tid, tname) in enumerate(targets):
                if tid.startswith('monitor:'):
                    display_text = f"🖥️ {tname}"
                elif tid == 'all':
                    display_text = "📺 All Screens"
                elif tid == 'disabled':
                    display_text = "❌ Disabled"
                else:
                    display_text = tname
                self.screenshot_combo.append_text(display_text)
            
            try:
                default_index = self._screenshot_target_map.index(self.screenshot_target)
            except ValueError:
                try:
                    default_index = self._screenshot_target_map.index('all')
                except ValueError:
                    default_index = 0
            self.screenshot_combo.set_active(default_index)
            self.screenshot_combo.connect("changed", self._on_screenshot_target_changed)
            
            refresh_btn = Gtk.Button(label="Refresh")
            refresh_btn.connect("clicked", self._on_refresh_screens_clicked)
            grid.attach(refresh_btn, 2, row, 1, 1)
        except Exception as e:
            print(f"Fehler beim Laden der Screenshot-Targets: {e}")
        
        grid.attach(l_screenshot, 0, row, 1, 1)
        grid.attach(self.screenshot_combo, 1, row, 1, 1)
        grid.attach(self.screenshot_disable_btn, 3, row, 1, 1)
    
    def _create_screenshot_interval_ui(self, grid, row):
        """Create screenshot interval UI."""
        l_interval = Gtk.Label(label="Screenshot Interval (s):")
        l_interval.set_xalign(0)
        adj = Gtk.Adjustment(value=self.screenshot_interval, lower=1, upper=300, step_increment=1)
        self.screenshot_interval_spin = Gtk.SpinButton(adjustment=adj)
        self.screenshot_interval_spin.connect("value-changed", self._on_screenshot_interval_changed)

        grid.attach(l_interval, 0, row, 1, 1)
        grid.attach(self.screenshot_interval_spin, 1, row, 1, 1)

    def _on_screenshot_interval_changed(self, widget):
        """Handle screenshot interval changes."""
        self.screenshot_interval = int(widget.get_value())
        settings = load_user_settings()
        settings['screenshot_interval'] = self.screenshot_interval
        save_user_settings(settings)
        print(f"Screenshot interval set to: {self.screenshot_interval} seconds")
    
    def _create_send_screenshots_to_llm_ui(self, grid, row):
        """Create UI for enabling/disabling sending screenshots to LLM."""
        self.send_screenshots_to_llm_check = Gtk.CheckButton(label="Send screenshots to LLM for analysis")
        self.send_screenshots_to_llm_check.set_active(bool(self.send_screenshots_to_llm))
        self.send_screenshots_to_llm_check.connect("toggled", self._on_send_screenshots_to_llm_toggled)
        
        grid.attach(self.send_screenshots_to_llm_check, 0, row, 2, 1)

    def _on_send_screenshots_to_llm_toggled(self, widget):
        """Handle toggling sending screenshots to LLM setting."""
        enabled = bool(widget.get_active())
        self.send_screenshots_to_llm = enabled
        
        settings = load_user_settings()
        settings['send_screenshots_to_llm'] = enabled
        save_user_settings(settings)
        
        print(f"Sending screenshots to LLM set to: {enabled}")

    def _create_consistency_check_ui(self, grid, row):
        """Create UI for enabling/disabling the consistency check."""
        self.consistency_check_check = Gtk.CheckButton(label="Consistency check")
        self.consistency_check_check.set_active(self.consistency_check_enabled)
        self.consistency_check_check.connect("toggled", self._on_consistency_check_toggled)
        
        grid.attach(self.consistency_check_check, 2, row, 2, 1)

    def _on_consistency_check_toggled(self, widget):
        """Handle toggling consistency check setting."""
        enabled = bool(widget.get_active())
        self.consistency_check_enabled = enabled
        
        settings = load_user_settings()
        settings['consistency_check_enabled'] = enabled
        save_user_settings(settings)
        
        print(f"Consistency check set to: {enabled}")

        if self.consistency_check_enabled and self.is_recording and not self.consistency_check_thread:
            print("Starting consistency check thread...")
            self.consistency_check_stop_event.clear()
            self.consistency_check_thread = threading.Thread(target=self._run_consistency_check_loop, daemon=True)
            self.consistency_check_thread.start()
        elif (not self.consistency_check_enabled or not self.is_recording) and self.consistency_check_thread:
            print("Stopping consistency check thread...")
            self.consistency_check_stop_event.set()
            self.consistency_check_thread = None
    
    def _create_speaker_controls(self, grid, row):
        """Create speaker count controls."""
        l_min_speakers = Gtk.Label(label="Min. Sprecher:")
        l_min_speakers.set_xalign(0)
        adj_min = Gtk.Adjustment(value=0, lower=0, upper=20, step_increment=1)
        self.min_speakers_spin = Gtk.SpinButton(adjustment=adj_min)
        self.min_speakers_spin.connect("value-changed", self._on_speakers_changed)
        
        l_max_speakers = Gtk.Label(label="Max. Sprecher:")
        l_max_speakers.set_xalign(0)
        adj_max = Gtk.Adjustment(value=0, lower=0, upper=20, step_increment=1)
        self.max_speakers_spin = Gtk.SpinButton(adjustment=adj_max)
        self.max_speakers_spin.connect("value-changed", self._on_speakers_changed)
        
        speaker_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=10)
        speaker_box.pack_start(l_min_speakers, False, False, 0)
        speaker_box.pack_start(self.min_speakers_spin, True, True, 0)
        speaker_box.pack_start(l_max_speakers, False, False, 10)
        speaker_box.pack_start(self.max_speakers_spin, True, True, 0)
        grid.attach(speaker_box, 0, row, 3, 1)
    
    def _create_language_ui(self, grid, row):
        """Create language selection UI."""
        l_lang = Gtk.Label(label="Transcription Language:")
        l_lang.set_xalign(0)
        self.lang_combo = Gtk.ComboBoxText()
        self.lang_combo.set_size_request(400, -1)
        
        for label, code in self._all_lang_choices:
            self.lang_combo.append_text(label)
        
        self.lang_search = Gtk.Entry()
        self.lang_search.set_placeholder_text("Filter...")
        self.lang_search.set_size_request(80, -1)
        self.lang_search.connect("changed", self._on_lang_filter_changed)
        
        try:
            settings = load_user_settings()
            current_lang = settings.get('transcription_language', 'en')
            try:
                idx = self._lang_code_map.index(current_lang)
            except ValueError:
                idx = 1  # default to English
            self.lang_combo.set_active(idx)
        except Exception:
            self.lang_combo.set_active(1)
        
        self.lang_combo_handler_id = self.lang_combo.connect("changed", self._on_transcription_language_changed)
        grid.attach(l_lang, 0, row, 1, 1)
        grid.attach(self.lang_combo, 1, row, 1, 1)
        grid.attach(self.lang_search, 2, row, 1, 1)
    
    def _create_ai_ui(self, grid, row):
        """Create AI-related UI."""
        self.ai_record_check = Gtk.CheckButton(label="AI: Record meeting")
        try:
            self.ai_record_check.set_active(bool(self.ai_record_meeting))
        except Exception:
            self.ai_record_check.set_active(True)
        self.ai_record_check.connect("toggled", self._on_ai_record_toggled)
        grid.attach(self.ai_record_check, 3, row, 1, 1)
    
    def _create_ollama_ui(self, grid, row):
        """Create Ollama URL setting UI."""
        l_ollama = Gtk.Label(label="Ollama URL:")
        l_ollama.set_xalign(0)
        self.ollama_entry = Gtk.Entry()
        try:
            settings = load_user_settings()
            self.ollama_entry.set_text(settings.get('ollama_url', 'http://localhost:11434'))
        except Exception:
            self.ollama_entry.set_text('http://localhost:11434')
        self.ollama_entry.set_size_request(400, -1)
        self.ollama_entry.connect("changed", self._on_ollama_url_changed)
        grid.attach(l_ollama, 0, row, 1, 1)
        grid.attach(self.ollama_entry, 1, row, 2, 1)

        # Ollama Model Name setting
        l_ollama_model = Gtk.Label(label="Ollama Model Name:")
        l_ollama_model.set_xalign(0)
        self.ollama_model_entry = Gtk.Entry()
        try:
            settings = load_user_settings()
            self.ollama_model_entry.set_text(settings.get('ollama_model_name', 'llama3'))
        except Exception:
            self.ollama_model_entry.set_text('llama3')
        self.ollama_model_entry.set_size_request(400, -1)
        self.ollama_model_entry.connect("changed", self._on_ollama_model_name_changed)
        grid.attach(l_ollama_model, 0, row + 1, 1, 1)
        grid.attach(self.ollama_model_entry, 1, row + 1, 2, 1)
    
    def _create_hf_token_ui(self, grid, row):
        """Create Hugging Face token UI."""
        l_hf = Gtk.Label(label="Hugging Face Token:")
        l_hf.set_xalign(0)
        self.hf_entry = Gtk.Entry()
        try:
            settings = load_user_settings()
            hf_token = settings.get('hf_token', cfg.HF_TOKEN or '')
            self.hf_entry.set_text(hf_token or '')
        except Exception:
            self.hf_entry.set_text(cfg.HF_TOKEN or '')
        self.hf_entry.set_size_request(400, -1)
        self.hf_entry.connect("changed", self._on_hf_token_changed)
        grid.attach(l_hf, 0, row, 1, 1)
        grid.attach(self.hf_entry, 1, row, 2, 1)
    
    def _create_joplin_ui(self, grid, row):
        """Create Joplin-related UI."""
        # Row 1: API Key
        l_joplin_key = Gtk.Label(label="Joplin API Key:")
        l_joplin_key.set_xalign(0)
        self.joplin_api_key_entry = Gtk.Entry()
        self.joplin_api_key_entry.set_text(self.joplin_api_key)
        self.joplin_api_key_entry.set_size_request(400, -1)
        self.joplin_api_key_entry.connect("changed", self._on_joplin_api_key_changed)
        grid.attach(l_joplin_key, 0, row, 1, 1)
        grid.attach(self.joplin_api_key_entry, 1, row, 2, 1)

        # Row 2: Destination Folder
        l_joplin_dest = Gtk.Label(label="Joplin Folder:")
        l_joplin_dest.set_xalign(0)
        self.joplin_destination_folder_entry = Gtk.Entry()
        self.joplin_destination_folder_entry.set_text(self.joplin_destination_folder)
        self.joplin_destination_folder_entry.set_size_request(400, -1)
        self.joplin_destination_folder_entry.connect("changed", self._on_joplin_destination_folder_changed)
        grid.attach(l_joplin_dest, 0, row + 1, 1, 1)
        grid.attach(self.joplin_destination_folder_entry, 1, row + 1, 2, 1)
        
        # Row 3: Sync Checkbox
        self.joplin_sync_check = Gtk.CheckButton(label="Sync Meeting Notes to Joplin")
        self.joplin_sync_check.set_active(self.joplin_sync_enabled)
        self.joplin_sync_check.connect("toggled", self._on_joplin_sync_toggled)
        grid.attach(self.joplin_sync_check, 1, row + 2, 2, 1)
    
    def _create_folder_ui(self, grid, row):
        """Create folder management UI."""
        l_folders = Gtk.Label(label="Source Folders:")
        l_folders.set_xalign(0)
        self.folders_listbox = Gtk.ListBox()
        self.folders_listbox.set_selection_mode(Gtk.SelectionMode.SINGLE)
        
        for folder in self.source_folders:
            label = Gtk.Label(label=folder)
            label.set_xalign(0)
            self.folders_listbox.add(label)
        
        scrolled_win = Gtk.ScrolledWindow()
        scrolled_win.set_policy(Gtk.PolicyType.NEVER, Gtk.PolicyType.AUTOMATIC)
        scrolled_win.add(self.folders_listbox)
        scrolled_win.set_min_content_height(100)
        
        add_folder_btn = Gtk.Button(label="Add Folder")
        add_folder_btn.connect("clicked", self.folder_manager.on_add_folder_clicked)
        
        remove_folder_btn = Gtk.Button(label="Remove")
        remove_folder_btn.connect("clicked", self.folder_manager.on_remove_folder_clicked)
        
        folder_button_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        folder_button_box.pack_start(add_folder_btn, False, False, 0)
        folder_button_box.pack_start(remove_folder_btn, False, False, 0)
        
        folder_vbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=6)
        folder_vbox.pack_start(scrolled_win, True, True, 0)
        folder_vbox.pack_start(folder_button_box, False, False, 0)
        
        grid.attach(l_folders, 0, row, 1, 1)
        grid.attach(folder_vbox, 1, row, 2, 1)
    
    def _create_timer_ui(self, grid, row):
        """Create timer UI."""
        l_timer = Gtk.Label(label="Recording Timer:")
        l_timer.set_xalign(0)
        self.timer_entry = Gtk.Entry()
        self.timer_entry.set_text(self.timer_setting_str)
        self.timer_entry.set_placeholder_text("hh:mm:ss")
        self.timer_entry_handler_id = self.timer_entry.connect("changed", self.timer_manager.on_timer_changed)
        
        self.timer_enable_btn = Gtk.CheckButton(label="Timer enabled")
        self.timer_enable_btn.set_active(self.timer_enabled)
        self.timer_enable_btn.connect("toggled", self.timer_manager.on_timer_enabled_toggled)
        
        timer_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=10)
        timer_box.pack_start(self.timer_entry, True, True, 0)
        timer_box.pack_start(self.timer_enable_btn, False, False, 0)
        
        grid.attach(l_timer, 0, row, 1, 1)
        grid.attach(timer_box, 1, row, 2, 1)
    
    def _create_action_buttons(self, grid, row):
        """Create action buttons."""
        self._record_button = Gtk.Button(label="Start recording")
        self._record_button.connect("clicked", self._on_record_button_clicked)
        self._update_record_button_label()
        
        test_screenshot_btn = Gtk.Button(label="Take Screenshot")
        test_screenshot_btn.connect("clicked", self._on_test_screenshot_clicked)
        
        button_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=10)
        button_box.set_halign(Gtk.Align.CENTER)
        button_box.pack_start(self._record_button, True, True, 0)
        button_box.pack_start(test_screenshot_btn, True, True, 0)
        
        self.meeting_browser_btn = Gtk.Button(label="Meeting Browser")
        self.meeting_browser_btn.connect("clicked", self._on_meeting_browser_clicked)
        button_box.pack_start(self.meeting_browser_btn, True, True, 0)

        self.chat_browser_btn = Gtk.Button(label="Chat Browser")
        self.chat_browser_btn.connect("clicked", self._on_chat_browser_clicked)
        button_box.pack_start(self.chat_browser_btn, True, True, 0)
        grid.attach(button_box, 0, row, 3, 1)
    
    def _on_meeting_browser_clicked(self, widget):
        """Callback for the 'Meeting Browser' button."""
        print("DEBUG: Meeting Browser button clicked.")
        if not self.meeting_browser_window:
            self.meeting_browser_window = MeetingBrowserWindow(self)
        GLib.idle_add(self.meeting_browser_window.create_or_show)

    def _on_chat_browser_clicked(self, widget):
        """Callback for the 'Chat Browser' button."""
        print("DEBUG: Chat Browser button clicked.")
        if not self.chat_browser_window:
            self.chat_browser_window = ChatBrowserWindow(self)
        GLib.idle_add(self.chat_browser_window.create_or_show)

    def reprocess_meeting(self, folder_name, callback=None):
        def reprocess_thread_func():
            print(f"Starting reprocessing for {folder_name}")
            
            now_str = folder_name.replace("meeting-", "")
            audio_file = os.path.join(folder_name, f"meeting-{now_str}-mixed.mp3")
            transcript_file = os.path.join(folder_name, f"meeting-{now_str}.txt")
            
            if not os.path.exists(audio_file):
                print(f"Audio file not found for reprocessing: {audio_file}")
                if callback:
                    GLib.idle_add(callback)
                return

            if os.path.exists(transcript_file):
                os.remove(transcript_file)
            
            file_lock = threading.Lock()
            
            if self.transcriber:
                self.transcriber.transcribe_recording_file(audio_file, transcript_file, file_lock)
                self.transcriber.analyze_meeting_minutes(transcript_file, self.send_screenshots_to_llm)
            else:
                 print("Transcriber not available.")

            print(f"Finished reprocessing for {folder_name}")
            if callback:
                GLib.idle_add(callback)

        thread = threading.Thread(target=reprocess_thread_func, daemon=True)
        thread.start()

    # ==================== Event Handlers ====================
    
    def _on_screenshot_disabled_toggled(self, widget):
        """Handle toggling screenshot disabled state."""
        try:
            self.screenshot_disabled = bool(widget.get_active())
        except Exception:
            self.screenshot_disabled = bool(widget.get_active())
        
        if self.screenshot_combo:
            self.screenshot_combo.set_sensitive(not self.screenshot_disabled)
        
        print(f"Screenshot disabled set to: {self.screenshot_disabled}")
        self._save_screenshot_settings()
    
    def _on_screenshot_target_changed(self, widget):
        """Handle user changing the screenshot target selection."""
        try:
            idx = widget.get_active()
            if idx is None:
                return
            self.screenshot_target = self._screenshot_target_map[int(idx)]
        except Exception:
            self.screenshot_target = 'full-screen'
        
        print(f"Screenshot target set to: {self.screenshot_target}")
        self._save_screenshot_settings()
    
    def _on_refresh_screens_clicked(self, widget):
        """Re-query available monitors and repopulate the screenshot combo."""
        try:
            targets = ScreenshotManager.list_capture_screens()
            prev = self.screenshot_target
            
            self.screenshot_combo.remove_all()
            self._screenshot_target_map = []
            for tid, tname in targets:
                if tid.startswith('monitor:'):
                    display_text = f"🖥️ {tname}"
                elif tid == 'all':
                    display_text = "📺 All Screens"
                elif tid == 'disabled':
                    display_text = "❌ Disabled"
                else:
                    display_text = tname
                self.screenshot_combo.append_text(display_text)
                self._screenshot_target_map.append(tid)
            
            try:
                idx = self._screenshot_target_map.index(prev)
            except ValueError:
                try:
                    idx = self._screenshot_target_map.index('all')
                except ValueError:
                    idx = 0
            self.screenshot_combo.set_active(idx)
        except Exception as e:
            print(f"Fehler beim Aktualisieren der Monitorliste: {e}")
    
    def _on_test_screenshot_clicked(self, widget):
        """Test screenshot capture with current settings."""
        try:
            import tempfile
            tmpdir = tempfile.mkdtemp()
            
            test_mgr = ScreenshotManager(tmpdir)
            test_mgr.capture_target = self.screenshot_target
            test_mgr.disabled = self.screenshot_disabled
            
            print(f"\n📸 Test Screenshot: target={self.screenshot_target}, disabled={self.screenshot_disabled}")
            print(f"📸 Saving to: {tmpdir}")
            
            test_mgr._refresh_monitors()
            print(f"📸 Monitor map: {test_mgr.monitor_map}")
            
            test_mgr._take_screenshot()
            
            import os
            files = os.listdir(tmpdir)
            if files:
                for f in files:
                    fpath = os.path.join(tmpdir, f)
                    size = os.path.getsize(fpath)
                    print(f"✅ Screenshot created: {f} ({size} bytes)")
            else:
                print("❌ No screenshot files created")
                
        except Exception as e:
            print(f"❌ Test screenshot error: {e}")
            import traceback
            traceback.print_exc()
    
    def _on_transcription_language_changed(self, widget):
        """Handle transcription language changes from the settings UI."""
        try:
            idx = widget.get_active()
            if idx is None or idx < 0:
                return
            lang_code = self._lang_code_map[idx]
            
            # Persist the new setting
            settings = load_user_settings()
            settings['transcription_language'] = lang_code
            save_user_settings(settings)
            
            print(f"Transcription language set to: {lang_code}")
            
            # If transcriber is running, update its language dynamically
            try:
                if self.transcriber:
                    self.transcriber.transcription_language = lang_code
            except Exception:
                pass
        except Exception as e:
            print(f"Error changing transcription language: {e}")
        
        self._save_screenshot_settings(lang_code)
    
    def _save_screenshot_settings(self, lang_code=None):
        """Save screenshot and language settings."""
        try:
            settings = load_user_settings()
            settings['screenshot_target'] = self.screenshot_target
            settings['screenshot_disabled'] = bool(self.screenshot_disabled)
            if lang_code:
                settings['transcription_language'] = lang_code
            save_user_settings(settings)
        except Exception:
            pass
    
    def _on_ai_record_toggled(self, widget):
        """Handle toggling AI: Record meeting setting."""
        try:
            enabled = bool(widget.get_active())
            self.ai_record_meeting = enabled
            
            try:
                settings = load_user_settings()
                settings['ai_record_meeting'] = enabled
                save_user_settings(settings)
            except Exception:
                pass
            
            print(f"AI Record meeting set to: {enabled}")
            
            try:
                if self.transcriber:
                    setattr(self.transcriber, 'ai_record_meeting', enabled)
            except Exception:
                pass
        except Exception as e:
            print(f"Error toggling AI record setting: {e}")
    
    def _on_lang_filter_changed(self, widget):
        """Filter available languages in the combo based on the search entry."""
        try:
            text = widget.get_text().strip().lower()
            if not text:
                filtered = self._all_lang_choices
            else:
                filtered = [pair for pair in self._all_lang_choices 
                           if text in pair[0].lower() or text in pair[1].lower()]
                if not filtered:
                    filtered = self._all_lang_choices
            
            # remember current selection
            try:
                settings = load_user_settings()
                current = settings.get('transcription_language', None)
            except Exception:
                current = None
            
            # repopulate combo
            self.lang_combo.remove_all()
            self._lang_code_map = [code for _, code in filtered]
            for label, code in filtered:
                self.lang_combo.append_text(label)
            
            # restore selection if still present
            if current and current in self._lang_code_map:
                try:
                    idx = self._lang_code_map.index(current)
                    self.lang_combo.set_active(idx)
                except Exception:
                    pass
            
        except Exception as e:
            print(f"Error filtering languages: {e}")
    
    def _on_speakers_changed(self, widget):
        """Handle speaker count adjustment."""
        if self.transcriber:
            min_speakers = int(self.min_speakers_spin.get_value())
            max_speakers = int(self.max_speakers_spin.get_value())
            self.transcriber.min_speakers = min_speakers
            self.transcriber.max_speakers = max_speakers
            print(f"Diarization speaker hint updated: min={min_speakers}, max={max_speakers}")
    
    def _on_ollama_url_changed(self, widget):
        """Persist Ollama URL when user changes the entry and update runtime analyzer."""
        try:
            url = widget.get_text().strip()
        except Exception:
            return
        
        if not url:
            return
        
        try:
            settings = load_user_settings()
            settings['ollama_url'] = url
            save_user_settings(settings)
            print(f"Ollama URL set to: {url}")
        except Exception:
            pass
        
        # Update existing analyzer if present
        try:
            analyzer = getattr(self.chat_window, 'ollama_analyzer', None)
            if analyzer:
                analyzer.base_url = url
                analyzer.endpoint = f"{url.rstrip('/')}/api/generate"
                analyzer.model = self.ollama_model_name # Ensure model name is also updated
        except Exception:
            pass

    def _on_ollama_model_name_changed(self, widget):
        """Persist Ollama model name when user changes the entry and update runtime analyzer."""
        try:
            model_name = widget.get_text().strip()
        except Exception:
            return
        
        if not model_name:
            return
        
        try:
            settings = load_user_settings()
            settings['ollama_model_name'] = model_name
            save_user_settings(settings)
            print(f"Ollama model name set to: {model_name}")
            self.ollama_model_name = model_name # Update instance variable
        except Exception:
            pass
        
        # Update existing analyzer if present
        try:
            analyzer = getattr(self.chat_window, 'ollama_analyzer', None)
            if analyzer:
                analyzer.model = model_name
        except Exception:
            pass

    def _on_whisper_model_changed(self, widget):
        """Persist Whisper model name when user changes the entry."""
        try:
            model_name = widget.get_text().strip()
        except Exception:
            return
        
        if not model_name:
            return
        
        try:
            settings = load_user_settings()
            settings['whisper_model_name'] = model_name
            save_user_settings(settings)
            print(f"Whisper model name set to: {model_name}")
            self.whisper_model_name = model_name # Update instance variable
        except Exception:
            pass
    
    def _on_hf_token_changed(self, widget):
        """Persist Hugging Face token when user changes the entry and update runtime config.HF_TOKEN."""
        try:
            token = widget.get_text().strip()
        except Exception:
            return
        
        # Save into user settings
        try:
            settings = load_user_settings()
            settings['hf_token'] = token
            save_user_settings(settings)
            print("Hugging Face token updated in settings")
        except Exception:
            pass
        
        # Also update the in-memory config variable used by other modules
        try:
            cfg.HF_TOKEN = token
        except Exception:
            pass
    
    def _on_joplin_api_key_changed(self, widget):
        """Persist Joplin API key when user changes the entry."""
        try:
            api_key = widget.get_text().strip()
            self.joplin_api_key = api_key
        except Exception:
            return
        
        try:
            settings = load_user_settings()
            settings['joplin_api_key'] = api_key
            save_user_settings(settings)
            print("Joplin API key updated in settings")
        except Exception as e:
            print(f"Error saving Joplin API key: {e}")
    
    def _on_joplin_sync_toggled(self, widget):
        """Handle toggling Joplin sync setting."""
        try:
            enabled = bool(widget.get_active())
            self.joplin_sync_enabled = enabled
            
            try:
                settings = load_user_settings()
                settings['joplin_sync_enabled'] = enabled
                save_user_settings(settings)
                print(f"Joplin sync enabled set to: {enabled}")
            except Exception as e:
                print(f"Error saving Joplin sync setting: {e}")
                
        except Exception as e:
            print(f"Error toggling Joplin sync setting: {e}")
            
    def _on_joplin_destination_folder_changed(self, widget):
        """Persist Joplin destination folder when user changes the entry."""
        try:
            destination_folder = widget.get_text().strip()
            self.joplin_destination_folder = destination_folder
        except Exception:
            return
        
        try:
            settings = load_user_settings()
            settings['joplin_destination_folder'] = destination_folder
            save_user_settings(settings)
            print(f"Joplin destination folder set to: {destination_folder}")
        except Exception as e:
            print(f"Error saving Joplin destination folder: {e}")
    
    # ==================== System Tray Callbacks ====================
    
    def _on_systray_record_clicked(self, icon, item):
        """Systray callback to start/stop recording."""
        GLib.idle_add(self._on_record_button_clicked, 
                     self._record_button if hasattr(self, '_record_button') else None)
    
    def _on_systray_screenshot_clicked(self, icon, item):
        """Systray callback for taking a manual screenshot."""
        if self.is_recording and self.screenshot_manager:
            print("📸 Manual screenshot triggered from systray.")
            thread = threading.Thread(target=self.screenshot_manager._take_screenshot, daemon=True)
            thread.start()
    
    def _on_systray_lang_selected(self, lang_code, *args):
        """Callback when a language is selected from the systray menu."""
        def update_logic():
            # 1. Save setting
            settings = load_user_settings()
            settings['transcription_language'] = lang_code
            save_user_settings(settings)
            print(f"Transcription language set to: {lang_code}")
            
            # 2. Update transcriber
            if self.transcriber:
                self.transcriber.transcription_language = lang_code
            
            # 3. Update settings window UI
            if self.lang_combo:
                try:
                    idx = self._lang_code_map.index(lang_code)
                    
                    if self.lang_combo_handler_id > 0:
                        self.lang_combo.handler_block(self.lang_combo_handler_id)
                    
                    self.lang_combo.set_active(idx)
                    
                    if self.lang_combo_handler_id > 0:
                        self.lang_combo.handler_unblock(self.lang_combo_handler_id)
                    
                except ValueError:
                    print(f"Note: Language '{lang_code}' not visible in filtered settings window.")
                except Exception as e:
                    print(f"Note: Could not visually update language in settings window. {e}")
            
            # 4. Rebuild systray menu
            self.system_tray.update_menu()
        
        GLib.idle_add(update_logic)
    
    def show_window(self, icon, item):
        """Callback to show window from system tray."""
        GLib.idle_add(self._create_or_show_window)
    
    def open_chat_window(self, icon, item):
        """Tray callback to open the chat window."""
        GLib.idle_add(self.chat_window.create_or_show)
    
    # ==================== Recording Management ====================
    
    def _start_file_recording(self):
        """Start recording transcriptions to file and create meeting folder with screenshots."""
        self.is_starting = True
        self._update_record_button_label()

        mic_name, loopback_name = None, None
        if self._window and self.mic_combo and self.mic_combo.get_active() != -1:
            mic_name = self.mic_combo.get_active_text()
            loopback_name = self.loopback_combo.get_active_text()
        else:
            mic_name = self.audio_device_manager.mic_dev_name
            loopback_name = self.audio_device_manager.loopback_dev_name

        if not mic_name or not loopback_name:
            print("ERROR: No audio devices selected or saved. Please open the settings window.")
            GLib.idle_add(self._create_or_show_window)
            self.is_starting = False
            self._update_record_button_label()
            return False

        print(f"Attempting to start recording with Mic: '{mic_name}' and Loopback: '{loopback_name}'...")

        mic_id, loopback_id = self.audio_device_manager.get_device_ids_from_names(mic_name, loopback_name)

        while mic_id is None or loopback_id is None:
            if not self.is_starting:  # Check if user cancelled
                print("Recording start cancelled by user.")
                self._update_record_button_label()
                return False

            print(f"Waiting for audio devices to become available... Checking again in 1 second.")
            
            # Process GTK events to keep UI responsive and allow cancellation
            for _ in range(10): # 10 * 0.1s = 1s
                if not self.is_starting:
                    break
                Gtk.main_iteration_do(blocking=False)
                time.sleep(0.1)

            mic_id, loopback_id = self.audio_device_manager.get_device_ids_from_names(mic_name, loopback_name)

        if not self.is_starting:
            print("Recording start cancelled by user.")
            self._update_record_button_label()
            return False

        self.is_starting = False
        
        # Ensure audio processing thread is running with the correct devices
        if (self.recorder is None or self.transcriber is None or
            self.mic_dev_id != mic_id or self.loopback_dev_id != loopback_id):
            self._start_audio_processing_thread(mic_id, loopback_id)
        
        # --- Final Check and Recording Start ---
        if self.recorder is None or self.transcriber is None:
            print("ERROR: Audio recorder is not ready. Please try again.")
            return False
        
        self.is_recording = True
        self.transcription_history = []
        now_str = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        
        if self.consistency_check_enabled and not self.consistency_check_thread:
            print("Starting consistency check thread due to recording start...")
            self.consistency_check_stop_event.clear()
            self.consistency_check_thread = threading.Thread(target=self._run_consistency_check_loop, daemon=True)
            self.consistency_check_thread.start()
        
        # Disable device selection during recording
        if hasattr(self, 'mic_combo') and self.mic_combo:
            self.mic_combo.set_sensitive(False)
        if hasattr(self, 'loopback_combo') and self.loopback_combo:
            self.loopback_combo.set_sensitive(False)
        if hasattr(self, 'audio_refresh_btn') and self.audio_refresh_btn:
            self.audio_refresh_btn.set_sensitive(False)
        
        if self.timer_enabled:
            self.timer_manager.start_timer()
        
        # Create meeting folder
        self.meeting_folder = f"meeting-{now_str}"
        try:
            os.makedirs(self.meeting_folder, exist_ok=True)
            print(f"📁 Erstelle Ordner für Aufnahme: {self.meeting_folder}")
            # Create initial meeting record in DB
            self.current_meeting_id = self.db.create_meeting(
                folder_name=self.meeting_folder,
                title=f"Meeting {now_str}",
                status="Recording"
            )
            if not self.current_meeting_id:
                print("ERROR: Failed to create meeting record in database.")
                self.is_recording = False
                self.meeting_folder = None
                return False
        except OSError as e:
            print(f"Fehler beim Erstellen des Ordners {self.meeting_folder}: {e}")
            self.is_recording = False
            self.meeting_folder = None
            return False
        
        # Create transcript file in the meeting folder
        filename = os.path.join(self.meeting_folder, f"meeting-{now_str}.txt")
        self.output_filename = filename
        try:
            self.output_file = open(filename, 'w', encoding='utf-8')
            print(f"Schreibe Aufnahme in Datei: {filename}")
        except IOError as e:
            print(f"Fehler beim Öffnen der Datei {filename}: {e}")
            self.is_recording = False
            self.output_filename = None
            return False
        
        # Start audio file recording in the meeting folder
        mixed_filename = os.path.join(self.meeting_folder, f"meeting-{now_str}-mixed.mp3")
        print(f"DEBUG: Attempting to start audio file writing to: {mixed_filename}")
        self.recorder.start_audio_file_writing(mixed_filename)
        print("DEBUG: `start_audio_file_writing` called on recorder.")
        
        # Tell transcriber to enable real-time transcription
        try:
            if self.transcriber:
                self.transcriber.recording_enabled = True
        except Exception:
            pass
        
        # Start screenshot capture with user-selected settings
        self.screenshot_manager = ScreenshotManager(self.meeting_folder)
        try:
            self.screenshot_manager.capture_target = self.screenshot_target
            self.screenshot_manager.disabled = self.screenshot_disabled
            self.screenshot_manager.interval = self.screenshot_interval
            print(f"📸 Screenshot settings applied: target={self.screenshot_target}, disabled={self.screenshot_disabled}, interval={self.screenshot_interval}s")
        except Exception as e:
            print(f"⚠️  Error applying screenshot settings: {e}")
        self.screenshot_manager.start_capture()
        
        self._update_record_button_label()
        
        # Show recording indicator window
        if not self.recording_window:
            self.recording_window = RecordingIndicatorWindow(self)
        self.recording_window.show_all()
        
        return True
    
    def _stop_file_recording(self):
        """Stop recording transcriptions to file and re-transcribe for accurate speaker detection."""
        if self.is_starting:
            self.is_starting = False
            return

        if self.consistency_check_thread:
            print("Stopping consistency check thread due to recording stop...")
            self.consistency_check_stop_event.set()
            self.consistency_check_thread = None

        self.is_recording = False
        
        # Re-enable device selection after recording
        if hasattr(self, 'mic_combo') and self.mic_combo:
            self.mic_combo.set_sensitive(True)
        if hasattr(self, 'loopback_combo') and self.loopback_combo:
            self.loopback_combo.set_sensitive(True)
        if hasattr(self, 'audio_refresh_btn') and self.audio_refresh_btn:
            self.audio_refresh_btn.set_sensitive(True)
        
        self.timer_manager.stop_timer()
        
        # Reset timer state and UI to the configured default
        settings = load_user_settings()
        self.timer_setting_str = settings.get('timer_value', '00:00:00')
        self.timer_manager.timer_seconds = (self.timer_manager.hms_to_seconds(
            self.timer_setting_str) or 0)
        self.timer_manager.remaining_seconds = self.timer_manager.timer_seconds
        GLib.idle_add(self.timer_manager.reset_timer_display)
        
        # Stop screenshot capture
        if self.screenshot_manager:
            self.screenshot_manager.stop_capture()
            self.screenshot_manager = None
            
        # Hide recording indicator window
        if self.recording_window:
            self.recording_window.hide()

        # Hide and clear history window
        if self.history_window:
            self.history_window.hide()
            self.history_window.clear()
        
        with self.file_lock:
            if self.output_file:
                print("Aufnahmedatei wird geschlossen.")
                self.output_file.close()
                self.output_file = None
        
        # Stop audio file recording and get the path to the saved file
        audio_file_path = None
        if self.recorder:
            print("DEBUG: Attempting to stop audio file writing.")
            audio_file_path = self.recorder.stop_audio_file_writing()
            print(f"DEBUG: `stop_audio_file_writing` returned: {audio_file_path}")
        else:
            print("DEBUG: `self.recorder` is None in `_stop_file_recording`. Cannot stop file writing.")
        
        # Disable real-time transcription immediately after stopping
        try:
            if self.transcriber:
                self.transcriber.recording_enabled = False
        except Exception:
            pass
        
        # Re-transcribe the full recording for accurate speaker detection
        if audio_file_path and self.transcriber and self.output_filename:
            print(f"\n🔄 Starte vollständige Transkription der Aufnahme...")
            try:
                import os
                if os.path.exists(self.output_filename):
                    os.remove(self.output_filename)
                    print(f"📝 Existierende Datei gelöscht: {self.output_filename}")
            except Exception as e:
                print(f"⚠️  Fehler beim Löschen der alten Datei: {e}")
            
            # Re-transcribe and save to the cleared file
            self.transcriber.transcribe_recording_file(audio_file_path, self.output_filename, self.file_lock)
            
            # Read the final transcript from the file
            final_transcript = ""
            if os.path.exists(self.output_filename):
                with open(self.output_filename, 'r', encoding='utf-8') as f:
                    final_transcript = f.read()

            final_inconsistencies_note = ""
            if self.consistency_check_enabled and final_transcript.strip():
                print("\n🔄 Performing final consistency check on full transcript...")
                analyzer = self._get_ollama_analyzer()
                previous_meeting_context = self._get_relevant_past_meeting_context(final_transcript)
                if previous_meeting_context.strip():
                    inconsistencies = analyzer.find_inconsistencies(final_transcript, previous_meeting_context)
                    if inconsistencies.strip():
                        final_inconsistencies_note = f"\n[FINAL CONSISTENCY CHECK NOTE]: {inconsistencies}\n"
                        print(final_inconsistencies_note)
                        try:
                            with open(self.output_filename, 'a', encoding='utf-8') as f:
                                f.write(final_inconsistencies_note)
                        except IOError as e:
                            print(f"ERROR: Failed to write final inconsistency note to file: {e}")
                    else:
                        print("✅ No final inconsistencies found.")
                else:
                    print("⚠️  No past meeting context available for final inconsistency check.")

            # Suggest a title based on the transcript
            suggested_title = f"Meeting {datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}"
            if final_transcript:
                analyzer = self._get_ollama_analyzer()
                llm_title = analyzer.suggest_title(final_transcript)
                if llm_title:
                    suggested_title = llm_title
            
            # Ask the user for the final title
            final_title = self._ask_for_meeting_title(suggested_title)

            # Update DB with full transcript, title, and status
            if self.current_meeting_id and self.meeting_folder:
                self.db.update_meeting(
                    folder_name=self.meeting_folder,
                    title=final_title,
                    transcript=final_transcript + (final_inconsistencies_note if final_inconsistencies_note else ""),
                    status="Transcribed"
                )
                print(f"✅ Meeting '{self.meeting_folder}' updated with title and full transcript.")
            
            # Start analysis in a separate thread to not block the UI
            analysis_thread = threading.Thread(
                target=self.transcriber.analyze_meeting_minutes,
                args=(self.output_filename, self.send_screenshots_to_llm, final_inconsistencies_note,),
                daemon=True
            )
            analysis_thread.start()
        
        self._update_record_button_label()

    def _ask_for_meeting_title(self, default_title):
        """Shows a dialog to ask the user for a meeting title."""
        dialog = Gtk.MessageDialog(
            transient_for=self._window,
            flags=0,
            message_type=Gtk.MessageType.QUESTION,
            buttons=Gtk.ButtonsType.OK_CANCEL,
            text="Enter Meeting Title",
        )
        dialog.format_secondary_text("Please enter a title for the meeting. An AI-generated title has been suggested for you.")

        entry = Gtk.Entry()
        entry.set_text(default_title)
        entry.set_activates_default(True) # Allow Enter key to confirm

        # Add the entry to the dialog's content area
        content_area = dialog.get_content_area()
        content_area.pack_end(entry, True, True, 0)
        dialog.show_all()
        
        # Set the OK button as the default response
        ok_button = dialog.get_widget_for_response(Gtk.ResponseType.OK)
        if ok_button:
            ok_button.set_can_default(True)
            ok_button.grab_default()

        response = dialog.run()
        title = entry.get_text().strip()
        dialog.destroy()

        if response == Gtk.ResponseType.OK and title:
            return title
        else:
            return default_title
    
    def _add_note_to_meeting_minutes(self, note: str):
        """Appends a note to the current meeting's transcript file and updates the database."""
        if not self.is_recording or not self.output_file or not self.meeting_folder:
            print("WARNING: Cannot add note. Recording not active or output file not open.")
            return

        print(f"Adding note to meeting minutes: {note.strip()}")
        with self.file_lock:
            try:
                self.output_file.write(f"\n{note}\n")
                self.output_file.flush()
                
                # Update DB with the new content
                if self.current_meeting_id:
                    # Re-read the file to get the full updated transcript
                    with open(self.output_filename, 'r', encoding='utf-8') as f:
                        updated_transcript = f.read()
                    self.db.update_meeting(
                        folder_name=self.meeting_folder,
                        transcript=updated_transcript
                    )
            except IOError as e:
                print(f"ERROR: Failed to write note to file or update DB: {e}")

    def _update_record_button_label(self):
        """Update the record button's label based on the current recording state."""
        if self._record_button:
            if self.is_starting:
                self._record_button.set_label("Starting...")
                self._record_button.set_sensitive(True) # Allow cancelling
            elif self.is_recording:
                self._record_button.set_label("Stop recording")
                self._record_button.set_sensitive(True)
            else:
                self._record_button.set_label("Start recording")
                self._record_button.set_sensitive(True)
    
    def _on_record_button_clicked(self, widget=None):
        """Handle record button toggle from UI or systray."""
        if self.is_starting:
            self.is_starting = False
            # The loop in _start_file_recording will see this and exit.
            print("Recording start cancelled.")
            # No need to call _update_record_button_label here, 
            # it will be called when _start_file_recording returns
            return

        if self.is_recording:
            self._stop_file_recording()
        else:
            if self._start_file_recording():
                pass
        
        self._update_record_button_label()
        self.system_tray.update_menu()
    
    # ==================== Transcription and Analysis ====================
    
    def _on_new_transcription(self, text, source, speaker):
        """Callback when new transcription is available."""
        if not self.is_recording:
            return

        now = datetime.now()
        now_str = now.strftime('%Y-%m-%d %H:%M:%S')
        speaker_str = f"/{speaker}" if speaker else ""
        line = f"[{now_str}] [{source}{speaker_str}]: {text}"
        
        self.transcription_history.append(line)
        
        # Add to recent transcriptions deque with timestamp
        self.recent_transcriptions.append((now, line))
        
        # Trim old entries from recent_transcriptions (older than 5 minutes)
        five_minutes_ago = now - timedelta(minutes=5)
        while self.recent_transcriptions and self.recent_transcriptions[0][0] < five_minutes_ago:
            self.recent_transcriptions.popleft()


        if self.history_window and self.history_window.is_visible():
            GLib.idle_add(self.history_window.append_text, line)

        if not self.output_file:
            return
        
        with self.file_lock:
            if self.is_recording and self.output_file:
                file_line = f"{line}\n"
                try:
                    self.output_file.write(file_line)
                    self.output_file.flush()
                except IOError as e:
                    print(f"Fehler beim Schreiben in die Datei: {e}")
                    GLib.idle_add(self._on_record_button_clicked, self._record_button)
    
    def _on_analysis_complete(self, analysis, final_inconsistencies_note: str = ""):
        """Callback when meeting minutes analysis is complete."""
        print("\n✅ Meeting minutes analysis complete and saved!")
        
        if self.current_meeting_id and self.meeting_folder:
            self.db.update_meeting(
                folder_name=self.meeting_folder,
                analysis=analysis,
                status="Analyzed"
            )
            print(f"✅ Meeting '{self.meeting_folder}' updated with analysis.")

        self.joplin_sync.sync_analysis(analysis, self.meeting_folder, final_inconsistencies_note)

    def _on_level_update(self, source, rms_level):
        """Callback from AudioRecorder with new audio level for a source."""
        GLib.idle_add(self._update_level_indicators, source, rms_level)

    def _update_level_indicators(self, source, rms_level):
        """Update the level bar indicators in the UI."""
        if source == "LOOPBACK":
            if self.loopback_level_bar:
                self.loopback_level_bar.set_value(rms_level)
        
        if self.recording_window and self.recording_window.is_visible():
            self.recording_window.update_level(source, rms_level)
            
        return False # Important for GLib.idle_add
    
    # ==================== ChromaDB and File Upload ====================
    
    def _get_relevant_past_meeting_context(self, query: str) -> str:
        """
        Retrieves relevant context from past meetings based on a query.
        Uses ChromaDB for semantic search.
        """
        context = []
        if not hasattr(self, 'chroma_collection') or self.chroma_collection is None:
            print("WARNING: ChromaDB collection not initialized. Cannot retrieve past meeting context.")
            return ""

        try:
            # Query ChromaDB for relevant documents (e.g., past meeting summaries or decisions)
            # The query_text will be the current meeting's recent history
            docs = self._get_chroma_context(query_text=query, n_results=5) # Get top 5 relevant documents
            
            if docs:
                context.append("--- CONTEXT FROM PREVIOUS MEETINGS ---")
                for i, doc in enumerate(docs):
                    context.append(f"Document {i+1}:\n{doc}\n")
                context.append("--- END CONTEXT ---")
            else:
                print("DEBUG: No relevant past meeting context found in ChromaDB.")

        except Exception as e:
            print(f"ERROR retrieving past meeting context from ChromaDB: {e}")
        
        return "\n".join(context)
        
    def _get_chroma_context(self, query_text, n_results=3):
        """Retrieve short context documents from ChromaDB collection if available."""
        docs = []
        coll = getattr(self, 'chroma_collection', None)
        if not coll:
            return docs
        
        try:
            # Try common query signature
            try:
                res = coll.query(query_texts=[query_text], n_results=n_results)
            except TypeError:
                res = coll.query([query_text], n_results=n_results)
            
            # res expected to be a dict with 'documents'
            if isinstance(res, dict):
                documents = res.get('documents') or res.get('results')
                if documents and isinstance(documents, list):
                    # documents[0] for single-query
                    first = documents[0]
                    if isinstance(first, list):
                        docs = [d for d in first if d]
                    else:
                        docs = [d for d in documents if d]
            else:
                # Some clients may return a list-like structure
                try:
                    if isinstance(res, list) and res:
                        entry = res[0]
                        docs = entry.get('documents', []) if isinstance(entry, dict) else []
                except Exception:
                    docs = []
        except Exception:
            docs = []
        
        # Truncate long documents for prompt safety
        short_docs = []
        for d in docs:
            if not d:
                continue
            if len(d) > 1500:
                short_docs.append(d[:1500] + '...')
            else:
                short_docs.append(d)
        return short_docs
    
    def _process_uploaded_file(self, filepath):
        """Read a file, extract text content based on its type, and add it to the ChromaDB collection."""
        coll = getattr(self, 'chroma_collection', None)
        if not coll:
            GLib.idle_add(self.chat_window.append_text, "System", "Error: ChromaDB collection not available.")
            return
        
        content = ""
        error_message = None
        
        try:
            _, extension = os.path.splitext(filepath)
            extension = extension.lower()
            
            if extension in ['.txt', '.md']:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
            elif extension in ['.htm', '.html']:
                from bs4 import BeautifulSoup
                with open(filepath, 'r', encoding='utf-8') as f:
                    soup = BeautifulSoup(f, 'html.parser')
                    content = soup.get_text(separator='\n', strip=True)
            elif extension == '.docx':
                import docx
                doc = docx.Document(filepath)
                full_text = []
                for para in doc.paragraphs:
                    full_text.append(para.text)
                content = '\n'.join(full_text)
            elif extension == '.xlsx':
                import openpyxl
                workbook = openpyxl.load_workbook(filepath)
                full_text = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        row_text = []
                        for cell in row:
                            if cell.value:
                                row_text.append(str(cell.value))
                        if row_text:
                            full_text.append(' '.join(row_text))
                content = '\n'.join(full_text)
            elif extension == '.pdf':
                from PyPDF2 import PdfReader
                reader = PdfReader(filepath)
                full_text = []
                for page in reader.pages:
                    extracted_text = page.extract_text()
                    if extracted_text:
                        full_text.append(extracted_text)
                content = '\n'.join(full_text)
            else:
                error_message = f"Unsupported file type: {extension}. Please upload TXT, MD, HTML, HTM, DOCX, XLSX, or PDF."
            
            if error_message:
                GLib.idle_add(self.chat_window.append_text, "System", error_message)
                return
            
            if not content.strip():
                error_message = f"No text content found in file '{os.path.basename(filepath)}'."
                GLib.idle_add(self.chat_window.append_text, "System", error_message)
                return
            
            # Need a unique ID for the document
            doc_id = f"upload-{os.path.basename(filepath)}-{datetime.now().timestamp()}"
            
            # Metadata can be useful for filtering later
            metadata = {
                "source": "file_upload",
                "filename": os.path.basename(filepath),
                "upload_time": datetime.now().isoformat()
            }
            
            coll.add(
                documents=[content],
                metadatas=[metadata],
                ids=[doc_id]
            )
            
            GLib.idle_add(self.chat_window.append_text, "System", 
                         f"Successfully uploaded and indexed '{os.path.basename(filepath)}'.")
            
        except Exception as e:
            error_message = f"Failed to process file '{os.path.basename(filepath)}': {e}"
            print(error_message)
            import traceback
            traceback.print_exc()
            GLib.idle_add(self.chat_window.append_text, "System", error_message)
    
    # ==================== Application Lifecycle ====================
    
    def _get_ollama_analyzer(self):
        # This helper ensures we have an analyzer instance with current settings
        from config import load_user_settings
        settings = load_user_settings()
        base_url = settings.get('ollama_url', 'http://localhost:11434')
        model_name = settings.get('ollama_model_name', 'llama3')
        lang = 'en' # default
        if self.transcriber:
            lang = self.transcriber.transcription_language

        return OllamaAnalyzer(base_url=base_url, model_name=model_name, language=lang)

    def get_response_suggestion(self):
        history = "\n".join(self.transcription_history)
        if not history.strip():
            print("No transcription history to get a suggestion for.")
            return

        def worker():
            analyzer = self._get_ollama_analyzer()
            result = analyzer.get_suggestion(history, mode="respond")
            if result['success']:
                suggestion = f"\n[SUGGESTION (RESPOND)]:\n{result['response']}\n"
                print(suggestion)
                if self.history_window and self.history_window.is_visible():
                    GLib.idle_add(self.history_window.append_text, suggestion)
            else:
                error = f"\n[ERROR]: {result['error']}\n"
                print(error)
                if self.history_window and self.history_window.is_visible():
                    GLib.idle_add(self.history_window.append_text, error)

        threading.Thread(target=worker, daemon=True).start()

    def get_details_suggestion(self):
        history = "\n".join(self.transcription_history)
        if not history.strip():
            print("No transcription history to get details for.")
            return

        def worker():
            analyzer = self._get_ollama_analyzer()
            result = analyzer.get_suggestion(history, mode="details")
            if result['success']:
                suggestion = f"\n[SUGGESTION (DETAILS)]:\n{result['response']}\n"
                print(suggestion)
                if self.history_window and self.history_window.is_visible():
                    GLib.idle_add(self.history_window.append_text, suggestion)
            else:
                error = f"\n[ERROR]: {result['error']}\n"
                print(error)
                if self.history_window and self.history_window.is_visible():
                    GLib.idle_add(self.history_window.append_text, error)

        threading.Thread(target=worker, daemon=True).start()

    def get_summary_suggestion(self):
        # self.recent_transcriptions already contains only the last 5 minutes
        recent_history_lines = [item[1] for item in self.recent_transcriptions]
        history = "\n".join(recent_history_lines)
        if not history.strip():
            print("No recent transcription history to get a summary for.")
            return

        def worker():
            analyzer = self._get_ollama_analyzer()
            result = analyzer.get_summary(history)
            if result['success']:
                suggestion = f"\n[SUMMARY (LAST 5 MIN)]:\n{result['response']}\n"
                print(suggestion)
                if self.history_window and self.history_window.is_visible():
                    GLib.idle_add(self.history_window.append_text, suggestion)
            else:
                error = f"\n[ERROR]: {result['error']}\n"
                print(error)
                if self.history_window and self.history_window.is_visible():
                    GLib.idle_add(self.history_window.append_text, error)
        
        threading.Thread(target=worker, daemon=True).start()
    
    
    def quit_action(self, icon, item):
        """Handle application quit."""
        print("\n👋 Beendigung eingeleitet...")
        if self.is_recording:
            self._stop_file_recording()
        
        self.folder_manager.stop_folder_watcher()
        self.stop_threads()
        
        if self.db:
            self.db.close()

        # --- Free GPU Memory ---
        try:
            print("Freigabe der KI-Modelle aus dem Speicher...")
            del self.whisper_model
            del self.diarization_pipeline
            if torch.cuda.is_available():
                print("Leere CUDA-Cache...")
                torch.cuda.empty_cache()
            print("✅ Modelle freigegeben.")
        except Exception as e:
            print(f"⚠️ Fehler bei der Freigabe der Modelle: {e}")
        
        self.system_tray.stop()
        Gtk.main_quit()
    
    def run(self):
        """Start the application."""
        print("DEBUG: Application run() started.")

        # --- Load AI Models ---
        try:
            settings = load_user_settings()
            whisper_model_name = settings.get('whisper_model_name', 'base')
        except Exception:
            whisper_model_name = 'base'

        # --- Whisper Model Loading with Fallback ---
        device = "cuda" if torch.cuda.is_available() else "cpu"
        print(f"🧠 Lade Whisper-Modell ({whisper_model_name}) auf {device}...")
        try:
            # Try to load on GPU first if available
            self.whisper_model = whisper.load_model(whisper_model_name, device=device)
            if device == "cuda":
                print("✅ Whisper-Modell auf GPU geladen.")
        except torch.OutOfMemoryError:
            if device == "cuda":
                print("⚠️ CUDA out of memory. Fallback auf CPU für Whisper-Modell.")
                device = "cpu"
                self.whisper_model = whisper.load_model(whisper_model_name, device=device)
                print("✅ Whisper-Modell auf CPU geladen.")
            else:
                # OOM on CPU is a more critical problem
                print("❌ Out of memory beim Laden des Whisper-Modells auf CPU. Beende Anwendung.")
                raise # Re-raise the exception to stop the app
        except Exception as e:
            print(f"❌ Unerwarteter Fehler beim Laden des Whisper-Modells: {e}")
            raise
        
        print("\n--- Speaker Diarization Setup ---")
        self.diarization_pipeline = self._load_diarization_pipeline()
        
        # Ensure a chromadb directory exists for persistent storage
        try:
            chroma_dir = os.path.join(os.getcwd(), "chromadb")
            os.makedirs(chroma_dir, exist_ok=True)
            try:
                import chromadb
                # Use PersistentClient for actual file-based storage
                try:
                    self.chroma_client = chromadb.PersistentClient(path=chroma_dir)
                except Exception:
                    # Fallback for older chromadb versions
                    try:
                        from chromadb.config import Settings
                        settings = Settings(persist_directory=chroma_dir, chroma_db_impl="duckdb+parquet")
                        self.chroma_client = chromadb.Client(settings)
                    except Exception as e:
                        print(f"⚠️ Failed to initialize Chroma client: {e}")
                        self.chroma_client = None
                
                # Ensure a collection exists for meeting documents
                try:
                    self.chroma_collection = self.chroma_client.get_or_create_collection("recass_meetings")
                except Exception:
                    # older/newer chromadb versions may use create_collection
                    try:
                        self.chroma_collection = self.chroma_client.create_collection("recass_meetings")
                    except Exception:
                        self.chroma_collection = None
                
                try:
                    entries = os.listdir(chroma_dir)
                    print(f"✅ ChromaDB initialized at: {chroma_dir} (contents: {entries})")
                except Exception:
                    print(f"✅ ChromaDB initialized at: {chroma_dir}")
                
                if self.chroma_collection is not None:
                    self.folder_indexer = FolderIndexer(self.chroma_collection, self.db)
                    self.folder_manager.start_folder_watcher()
                    
            except Exception as e:
                print(f"⚠️ ChromaDB not available or failed to init: {e}")
                self.chroma_client = None
                self.chroma_collection = None
        except Exception as e:
            print(f"⚠️ Failed to prepare ChromaDB directory: {e}")
        
        # Start system tray
        self.system_tray.run()